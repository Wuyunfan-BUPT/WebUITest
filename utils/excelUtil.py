#!/usr/bin/env python 
# -*- coding: utf-8 -*-
# @Project : Guangfa UI test
# @Time    : 2022/6/1 19:18
# @Author  : wuyfee
# @File    : excelUtil.py
# @Software: PyCharm
import openpyxl
from config.conf import EXCEL_DIR
from utils.parseConfFile import ParseConfFile

do_conf = ParseConfFile()
dataName = do_conf.get_value('BaseSetting', 'dataName')

class ExcelUtil:


    def read_excel(self, tableName):
        wb = openpyxl.load_workbook(f"{EXCEL_DIR}/{dataName}")
        sheet = wb[tableName]

        all_items = []
        for row in range(2, sheet.max_row+1):
            items = []
            for col in range(2, sheet.max_column+1):
                items.append(sheet.cell(row, col).value)
            all_items.append(items)
        return all_items

if __name__ == '__main__':
    pass
    '''
    ExcelUtil().read_excel()
    '''